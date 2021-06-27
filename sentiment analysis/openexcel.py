from openpyxl import Workbook
from openpyxl import load_workbook

#在python中创建一个新的excel文件
wb = Workbook()
#每一个excel文件被创建时都自动生成一个工作表，通过以下指令进到该工作表
ws = wb.active
#创建新的工作表
ws1 = wb.create_sheet("工作表")
#通过工作表的名称为键，在文件中进到该工作表
ws2 = wb["工作表"]
#通过以下指令返还文件中的所有工作表
sheets = wb.sheetnames
#进入一个工作表后，通过以下指令对某一格中的值进行寻找或赋值，若该位置数据为空，则返还值为None
ws.cell(row = 1, column = 1).value = True
value = ws.cell(row = 1, column = 1).value
#保存该文件在本地电脑中
wb.save('/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/document.xlsx')

wb = load_workbook('/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/document.xlsx')

