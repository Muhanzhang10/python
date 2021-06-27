from openpyxl import Workbook
from openpyxl import load_workbook
import os 
from snownlp import SnowNLP
from snownlp import sentiment

paths = os.listdir(r'/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/output')
for path in paths:
    print(path)
    wb = load_workbook(r'/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/output/' + path)
    ws = wb['Sheet']
    ws.cell(row = 1, column = 5).value = '情感积分'
    for row in range(2, 10000000):
        if ws.cell(row = row, column = 1).value == None:
            break
        text = ws.cell(row = row, column = 2).value
        average = 0
        text_s = text.split('。')
        count = 0
        for i in range(len(text_s)):
            if text_s[i] == '':
                count += 1
                continue
            temp = SnowNLP(text_s[i])
            average += temp.sentiments
        average = average / (len(text_s) - count)
        ws.cell(row = row, column = 5).value = average
    wb.save(r'/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/output/' + path)
        
