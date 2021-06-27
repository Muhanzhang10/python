from openpyxl import Workbook
from openpyxl import load_workbook
import os 
from snownlp import SnowNLP
from snownlp import sentiment

'''
excel文件编辑： https://openpyxl.readthedocs.io/en/stable/tutorial.html
文件夹迭代： https://stackoverflow.com/questions/10377998/how-can-i-iterate-over-files-in-a-given-directory
忽略隐藏文件：https://stackoverflow.com/questions/45666120/os-listdir-is-giving-larger-lists-than-expected
snowNLP：https://medium.com/analytics-vidhya/python-snownlp-sentiment-analysis-for-the-chinese-language-8d9cafd0447

instructions
1 在每个文件中，抓取赞同数大于200或评论数大于20的评论，创建新的文件(存储在output文件夹中)，存储这些评论以及它们的点赞数和评论数。
2 对于里面评论依次进行分析：抓取重点三句话，对三句话依次进行情感分析，最后求平均，为该评论的情感值
3 对所有评论结果进行加权（根据点赞数量）取平均
'''
files = os.listdir(r'/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/data')

for file in files:
    pre = '/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/data'
    path = pre + '/' + file
    wb = load_workbook(path)
    name = wb.sheetnames[0]
    ws = wb[name]
    row = 2
    row2 = 2
    #创建一个新的excel文件
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.cell(row = 1, column = 1).value = "发布账号"
    ws2.cell(row = 1, column = 2).value = "正文"
    ws2.cell(row = 1, column = 3).value = "点赞数"
    ws2.cell(row = 1, column = 4).value = "评论数"
    while True:
        if ws.cell(row = row, column = 1).value == None:
            break
        if ws.cell(row=row, column = 5).value >= 200 or ws.cell(row=row, column = 6).value >= 20:
            ws2.cell(row = row2, column = 1).value = ws.cell(row =row, column = 1).value #发布账号
            ws2.cell(row = row2, column = 2).value = ws.cell(row =row, column = 3).value #正文
            ws2.cell(row = row2, column = 3).value = ws.cell(row =row, column = 5).value #点赞数
            ws2.cell(row = row2, column = 4).value = ws.cell(row =row, column = 6).value #评论数
            row2 += 1
        row += 1
    wb2.save('/Users/zhangmuhan/Desktop/杂项/Python课程/git/sentiment analysis/output/' + file)




        




        





